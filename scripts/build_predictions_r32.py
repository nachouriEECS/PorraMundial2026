#!/usr/bin/env python3
"""Build RondaEliminatoria/data/predictions.json from the Excel.

Parses the knockout blocks of `Porra Mundial 2026.xlsx`:
  * "Ronda Eliminatoria" — the 16 Round-of-32 matches (round "r32").
  * the 8 Round-of-16 matches that follow it (round "r16").
  * the 4 quarter-final matches that follow those (round "cf").
  * the 2 semifinal matches that follow those (round "sf").
  * the third-place play-off (round "3p") and the final (round "f"), in that
    row order — both sit in the "center" column rather than a bracket side.

Each player cell is an exact-score prediction like `4-1`, optionally annotated
with who advances on a draw (`2-2 (ALE)`, `1-1 CAN`, `1-1 pasa Canada`,
`2-2 (pen Brasil)`). Emits players, the matches in bracket order (with Spanish
display names + team codes, a round tag and side/pos for layout), and per-player
{g1, g2, adv}.

The Round-of-16 side/pos are derived from the Round-of-32 matches that feed each
tie, so the bracket lines the octavos up beside the pair of R32 boxes they come
from.

The trailing "Ganadores" row (who each player picked to lift the trophy) and the
"Preguntas" block (best goal, pichichi, own goals, most-conceded team, total
throw-ins) hold the six bonus questions. Those are emitted as `questions` — the
real answers live in data/bonus.json, scored in the browser.
"""
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "RondaEliminatoria" / "Porra Mundial 2026.xlsx"
OUT = ROOT / "RondaEliminatoria" / "data" / "predictions.json"

# Bracket order is the row order in the Excel. side/pos drive the layout:
# 8 matches down the left, 8 down the right, top -> bottom.
SIDES = (["left"] * 8) + (["right"] * 8)

# Spanish display name -> team code (codes match ../data/team-codes.json).
NAME2CODE = {
    "alemania": "GER", "paraguay": "PAR", "francia": "FRA", "suecia": "SWE",
    "sudafrica": "RSA", "canada": "CAN", "paisesbajos": "NED",
    "marruecos": "MAR", "portugal": "POR", "croacia": "CRO", "espana": "ESP",
    "austria": "AUT", "eeuu": "USA", "bosnia": "BOS", "belgica": "BEL",
    "senegal": "SEN", "brasil": "BRA", "japon": "JPN", "cdemarfil": "CIV",
    "cmarfil": "CIV", "noruega": "NOR", "mexico": "MEX", "ecuador": "ECU",
    "inglaterra": "ENG", "rdcongo": "CON", "argentina": "ARG",
    "caboverde": "CPV", "australia": "AUS", "egipto": "EGY", "suiza": "SUI",
    "argelia": "ALG", "colombia": "COL", "ghana": "GHA",
}
# Cells sometimes name a team by its 3-letter code (e.g. "USA - BEL").
VALID_CODES = set(NAME2CODE.values())

# Section headers that mark the end of the knockout prediction blocks.
STOP_HEADERS = {"puntos", "ganadores", "preguntas"}

# The tournament winner sits on its own "Ganadores" row above the "Preguntas"
# block, and unlike those it carries the answers on the header row itself.
WINNER = ("ganadores", "q-ganador", "Ganador del Mundial", 36, "team")

# Bonus questions, keyed by the normalized row label in the "Preguntas" block.
# `kind` drives how the browser scores them: "team" = exact answer, points split
# between everyone who nailed it; "number" = closest guess wins, split on a tie.
QUESTIONS = [
    ("mejorgol", "q-mejor-gol", "Selección con el mejor gol", 12, "team"),
    ("pichichi", "q-pichichi", "Selección del pichichi", 12, "team"),
    ("golesenpropia", "q-goles-propia", "Goles en propia", 12, "number"),
    ("masgoleada", "q-mas-goleada", "Selección más goleada", 10, "team"),
    ("banda", "q-saques-banda", "Saques de banda (total)", 10, "number"),
]


def norm(s):
    """Lowercase, strip accents and non-letters — for fuzzy name matching."""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z]", "", s.lower())


def name_to_code(name):
    raw = str(name).strip()
    if raw.upper() in VALID_CODES:      # a bare team code, e.g. "USA"
        return raw.upper()
    code = NAME2CODE.get(norm(name))
    if not code:
        sys.exit(f"Unknown team name: {name!r} (normalized {norm(name)!r})")
    return code


def split_teams(label):
    """Split a match label into its two team names.

    R32 rows use ` - ` (spaces), R16 rows use a bare `-`, so accept either.
    """
    parts = re.split(r"\s*-\s*", str(label).strip())
    if len(parts) != 2:
        sys.exit(f"Cannot split match label into two teams: {label!r}")
    return parts[0].strip(), parts[1].strip()


def resolve_advancer(token, t1, t2, es1, es2):
    """Pick which of the two teams an annotation token refers to.

    Scoped to the two teams in the row, so a prefix/code match is unambiguous.
    Filler words ("pen", "pasa") are dropped first; matches against the team code
    and the normalized Spanish name.
    """
    tok = norm(re.sub(r"pen|pasa", " ", token.lower()))
    if not tok:
        return None
    for code, es in ((t1, es1), (t2, es2)):
        n = norm(es)
        if tok == code.lower() or n.startswith(tok) or tok.startswith(code.lower()):
            return code
    return None


def parse_cell(raw, t1, t2, es1, es2, ctx):
    """Return {g1, g2, adv} or None for a blank cell."""
    if raw is None or str(raw).strip() == "":
        return None
    s = str(raw).strip()
    m = re.search(r"(\d+)\s*-\s*(\d+)", s)
    if not m:
        print(f"  WARN {ctx}: cannot parse score from {s!r}", file=sys.stderr)
        return None
    g1, g2 = int(m.group(1)), int(m.group(2))
    annotation = (s[: m.start()] + s[m.end():]).strip()
    if g1 > g2:
        adv = t1
    elif g2 > g1:
        adv = t2
    else:
        adv = resolve_advancer(annotation, t1, t2, es1, es2)
        if adv is None:
            print(f"  WARN {ctx}: drawn pick {s!r} has no advancer", file=sys.stderr)
    return {"g1": g1, "g2": g2, "adv": adv}


def find_header_row(ws, header):
    for r in range(1, ws.max_row + 1):
        if norm(ws.cell(r, 1).value or "") == header:
            return r
    return None


def read_players(ws):
    """Players come from row 1 (skip the "Partidos" corner cell)."""
    players, c = [], 2
    while ws.cell(1, c).value not in (None, ""):
        players.append(str(ws.cell(1, c).value).strip())
        c += 1
    return players


def read_block(ws, start_row, count, players, id_prefix, round_tag,
               matches, predictions, names, side_pos):
    """Read `count` match rows starting at `start_row`.

    `side_pos(idx, t1, t2)` returns the (side, pos) for a match; blank rows are
    skipped and a section header stops the block early. Returns the next row.
    """
    row, idx = start_row, 0
    while idx < count:
        label = ws.cell(row, 1).value
        if label is None or str(label).strip() == "":
            row += 1
            continue
        if norm(label) in STOP_HEADERS:
            break
        es1, es2 = split_teams(label)
        t1, t2 = name_to_code(es1), name_to_code(es2)
        side, pos = side_pos(idx, t1, t2)
        mid = f"{id_prefix}-{idx + 1:02d}"
        matches.append({"id": mid, "round": round_tag, "side": side, "pos": pos,
                        "team1": t1, "team2": t2})
        names.setdefault(t1, es1)
        names.setdefault(t2, es2)
        predictions[mid] = {}
        for i, player in enumerate(players):
            pick = parse_cell(ws.cell(row, 2 + i).value, t1, t2, es1, es2,
                              f"{mid} {player}")
            if pick is not None:
                predictions[mid][player] = pick
        idx += 1
        row += 1
    return row


def read_row_answers(ws, row, players):
    """Read one player-per-column answer row, keyed by the row-1 player names."""
    answers = {}
    for i, player in enumerate(players):
        val = ws.cell(row, 2 + i).value
        if val is None or str(val).strip() == "":
            continue
        answers[player] = str(val).strip()
    return answers


def read_winner(ws, players):
    """Read the "Ganadores" row — each player's pick to win the tournament."""
    key, qid, label, pts, kind = WINNER
    row = find_header_row(ws, key)
    if row is None:
        print("  WARN: no 'Ganadores' row found — skipping the winner question",
              file=sys.stderr)
        return []
    return [{"id": qid, "label": label, "points": pts, "kind": kind,
             "answers": read_row_answers(ws, row, players)}]


def read_questions(ws, players):
    """Read the bonus-question block that follows the bracket.

    The block repeats the player names in its own header row, but spelled
    inconsistently ("Nacho A" up top vs "Ignacio A" down here), so answers are
    taken by column position and keyed with the row-1 names instead.
    """
    header = find_header_row(ws, "preguntas")
    if header is None:
        print("  WARN: no 'Preguntas' block found — skipping bonus questions",
              file=sys.stderr)
        return []

    wanted = {key: (qid, label, pts, kind)
              for key, qid, label, pts, kind in QUESTIONS}
    found = {}
    for row in range(header + 1, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label is None or str(label).strip() == "":
            continue
        key = norm(label)
        if key not in wanted:
            continue
        found[key] = read_row_answers(ws, row, players)

    questions = []
    for key, qid, label, pts, kind in QUESTIONS:
        if key not in found:
            print(f"  WARN: bonus question {label!r} not found in the Excel",
                  file=sys.stderr)
            continue
        questions.append({"id": qid, "label": label, "points": pts,
                          "kind": kind, "answers": found[key]})
    return questions


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]

    header_row = find_header_row(ws, "rondaeliminatoria")
    if header_row is None:
        sys.exit("Could not find 'Ronda Eliminatoria' section in the Excel")

    players = read_players(ws)
    matches, predictions, names = [], {}, {}

    # --- Round of 32: 16 matches, laid out 8 left then 8 right, top -> bottom.
    next_row = read_block(
        ws, header_row + 1, 16, players, "r32", "r32",
        matches, predictions, names,
        side_pos=lambda idx, t1, t2: (SIDES[idx], idx % 8),
    )

    # --- Round of 16: 8 matches. Each tie is fed by two adjacent R32 matches
    #     (positions 2k and 2k+1 on one side), so derive its side/pos from them.
    team_r32 = {}
    for m in matches:
        team_r32[m["team1"]] = (m["side"], m["pos"])
        team_r32[m["team2"]] = (m["side"], m["pos"])

    def r16_side_pos(idx, t1, t2):
        s1, p1 = team_r32[t1]
        s2, p2 = team_r32[t2]
        if s1 != s2:
            sys.exit(f"R16 tie {t1}-{t2} draws from both bracket sides")
        return s1, min(p1, p2) // 2

    next_row = read_block(
        ws, next_row, 8, players, "r16", "r16",
        matches, predictions, names, side_pos=r16_side_pos,
    )

    # --- Quarter-finals: 4 matches. Each tie is fed by two adjacent R16 matches
    #     (positions 2k and 2k+1 on one side), so derive its side/pos from those
    #     the same way the R16 column was derived from the R32 boxes.
    team_r16 = {}
    for m in matches:
        if m["round"] != "r16":
            continue
        team_r16[m["team1"]] = (m["side"], m["pos"])
        team_r16[m["team2"]] = (m["side"], m["pos"])

    def cf_side_pos(idx, t1, t2):
        s1, p1 = team_r16[t1]
        s2, p2 = team_r16[t2]
        if s1 != s2:
            sys.exit(f"CF tie {t1}-{t2} draws from both bracket sides")
        return s1, min(p1, p2) // 2

    next_row = read_block(
        ws, next_row, 4, players, "cf", "cf",
        matches, predictions, names, side_pos=cf_side_pos,
    )

    # --- Semifinals: 2 matches. Each tie is fed by the pair of quarter-final
    #     matches on one bracket side, so derive its side/pos from those the same
    #     way the cuartos column was derived from the R16 boxes.
    team_cf = {}
    for m in matches:
        if m["round"] != "cf":
            continue
        team_cf[m["team1"]] = (m["side"], m["pos"])
        team_cf[m["team2"]] = (m["side"], m["pos"])

    def sf_side_pos(idx, t1, t2):
        s1, p1 = team_cf[t1]
        s2, p2 = team_cf[t2]
        if s1 != s2:
            sys.exit(f"SF tie {t1}-{t2} draws from both bracket sides")
        return s1, min(p1, p2) // 2

    next_row = read_block(
        ws, next_row, 2, players, "sf", "sf",
        matches, predictions, names, side_pos=sf_side_pos,
    )

    # --- Third place, then the final. Both draw from across the bracket, so
    #     they get the "center" column instead of a side; pos orders them within
    #     it (final on top, third-place play-off below the trophy card). Excel
    #     row order is the source of truth for which is which.
    next_row = read_block(
        ws, next_row, 1, players, "3p", "3p",
        matches, predictions, names, side_pos=lambda idx, t1, t2: ("center", 1),
    )
    read_block(
        ws, next_row, 1, players, "f", "f",
        matches, predictions, names, side_pos=lambda idx, t1, t2: ("center", 0),
    )

    # Winner first — it is the biggest pot, so it leads the bonus board.
    questions = read_winner(ws, players) + read_questions(ws, players)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(
        {"players": players, "matches": matches, "names": names,
         "predictions": predictions, "questions": questions},
        ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    counts = {}
    for m in matches:
        counts[m["round"]] = counts.get(m["round"], 0) + 1
    summary = ", ".join(f"{n} {r}" for r, n in counts.items())
    print(f"Wrote {OUT.relative_to(ROOT)}: {len(players)} players, "
          f"{len(matches)} matches ({summary})")


if __name__ == "__main__":
    main()
